import os
import openpyxl as xl
import openpyxl.comments

# Global variables
SUMMARY_SHEET_NAME = "Summary"
TARGET_MARKSHEET_NAME = "MarkEntry"
TARGET_FILE_FIRST_REPORT_COL = 5
TARGET_FILE_PROGRESS_COL = 6
TARGET_FILE_EXECUTIVE_PRESENTATION_COL = 7
TARGET_FILE_ORAL_COL = 8
TARGET_FILE_FINAL_REPORT_COL = 9

CELL_LOC_STUDENT_NUM = 'C7'
CELL_LOC_STUDENT_NAME = 'C6'
CELL_LOC_STUDENT_SUPERVISOR = "C10"
CELL_LOC_STUDENT_MODERATOR = 'C11'
CELL_LOC_STUDENT_PROGRESS = "H7"
CELL_LOC_STUDENT_FIRST_REPORT = 'H6'
CELL_LOC_STUDENT_EXEC = "H8"
CELL_LOC_STUDENT_ORAL = "H9"
CELL_LOC_STUDENT_FINAL_SUP = "E11"
CELL_LOC_STUDENT_FINAL_MOD = "E10"
CELL_LOC_STUDENT_FINAL = "H11"
CELL_LOC_STUDENT_TOTAL = "H12"


def load_id_number_list(roster_file_name, worksheet_name=TARGET_MARKSHEET_NAME):
    """
    Creates a list with the id numbers of the students.

    It opens the roster file, retrieves the id numbers of all students
    and returns them as a list.

    :param roster_file_name: target file
    :param worksheet_name: worksheet where the data are.
    If not specified the default value is "MarkEntry"
    :return: a list with all id numbers as strings
    """
    wb_january = xl.load_workbook(roster_file_name)
    id_number_list = []
    ws = wb_january[worksheet_name]
    i = 1
    offset = 5
    student_id = ws.cell(row=offset, column=1).value
    while student_id is not None:
        id_number_list += [str(student_id)]
        student_id = ws.cell(row=offset+i, column=1).value
        i += 1
    return id_number_list


def find_in_list(string_list, key_word):
    """
    Gives an index number for the given element in the given list.

    :param string_list: A list
    :param key_word: element to be found
    :return: the index number of the element
    :raises ValueError: if the element is not found
    """
    return string_list.index(key_word)


def fetch_marks_for_student(filename):
    """
    Creates a dictionary with all the marks and other info from a
    student's project marksheet.

    the dictionary contains the following keys:
    - student_number
    - student_name
    - supervisor: supervisor name
    - moderator: moderator name
    - first_report
    - progress
    - executive_presentation
    - oral
    - final_report_moderator
    - final_report_supervisor
    - final_report
    - total_mark

    :param filename: project marksheet filename
    :return: a dictionary
    """
    wb = xl.load_workbook(filename, data_only=True)
    ws = wb[SUMMARY_SHEET_NAME]
    student_number = str(ws[CELL_LOC_STUDENT_NUM].value)
    student_name = ws[CELL_LOC_STUDENT_NAME].value
    student_supervisor = ws[CELL_LOC_STUDENT_SUPERVISOR].value.strip()
    student_moderator = ws[CELL_LOC_STUDENT_MODERATOR].value.strip()
    first_report_mark = ws[CELL_LOC_STUDENT_FIRST_REPORT].value
    progress_mark = ws[CELL_LOC_STUDENT_PROGRESS].value
    executive_presentation_mark = ws[CELL_LOC_STUDENT_EXEC].value
    oral_mark = ws[CELL_LOC_STUDENT_ORAL].value
    final_report_moderator_mark = ws[CELL_LOC_STUDENT_FINAL_MOD].value
    final_report_supervisor_mark = ws[CELL_LOC_STUDENT_FINAL_SUP].value
    final_report = ws[CELL_LOC_STUDENT_FINAL].value
    total_mark = ws[CELL_LOC_STUDENT_TOTAL].value
    student_results = {
        "student_number": student_number,
        "student_name": student_name,
        "supervisor": student_supervisor,
        "moderator": student_moderator,
        "first_report": first_report_mark,
        "progress": progress_mark,
        "executive_presentation": executive_presentation_mark,
        "oral": oral_mark,
        "final_report_moderator": final_report_moderator_mark,
        "final_report_supervisor": final_report_supervisor_mark,
        "final_report": final_report,
        "total_mark": total_mark
    }
    return student_results


def record_data_for_student(results, roster_file_name, list_of_students_ids):
    """
    Writes the data into the target file.

    Finds the right position and fills in the data from a dictionary.
    It also inserts a comment with the name and the individual mark of
    the supervisor and the moderator

    :param results: dictionary with data
    :param roster_file_name: target file
    :param list_of_students_ids: list of all students numbers in their order
    of appearance in the target file
    :return: the target file filled with the data
    """
    wb = xl.open(roster_file_name)
    ws = wb[TARGET_MARKSHEET_NAME]
    loc = find_in_list(list_of_students_ids, results["student_number"])
    offset = 5
    ws.cell(row=loc+offset, column=TARGET_FILE_FIRST_REPORT_COL).value = results["first_report"]
    ws.cell(row=loc + offset, column=TARGET_FILE_PROGRESS_COL).value = results["progress"]
    ws.cell(row=loc + offset, column=TARGET_FILE_EXECUTIVE_PRESENTATION_COL).value = results["executive_presentation"]
    ws.cell(row=loc + offset, column=TARGET_FILE_ORAL_COL).value = results["oral"]
    ws.cell(row=loc + offset, column=TARGET_FILE_FINAL_REPORT_COL).value = results["final_report"]
    supervisor_mark = round(4 * results["final_report_supervisor"])
    moderator_mark = round(4 * results["final_report_moderator"])
    supervisor = results["supervisor"]
    moderator = results["moderator"]
    supervisor_and_moderator_marks = xl.comments.Comment(f"supervisor ({supervisor}): {supervisor_mark}, "
                                                         f"moderator ({moderator}): {moderator_mark}", "python")
    ws.cell(row=loc + offset, column=TARGET_FILE_FINAL_REPORT_COL).comment = supervisor_and_moderator_marks
    wb.save(roster_file_name)


def copy_markssheets_to_project_roster(marksheet_folder_name, roster_file_name):
    """
    Inserts the data for every student.

    It opens every file in the folder where the project marksheets are,
    finds the right position in the target file and copies them there

    :param marksheet_folder_name: folder where the marksheets are
    :param roster_file_name: target file
    """
    marksheets_list = os.listdir(marksheet_folder_name)
    student_numbers_in_roster = load_id_number_list(roster_file_name)
    for file in marksheets_list:
        # marksheet_folder_name = data/Mark Sheets Jan22
        # file = xyz.xlsx
        # what_we_want = data/Mark Sheets Jan22/xyz.xlsx
        file_path = os.path.join(marksheet_folder_name, file)
        marks = fetch_marks_for_student(file_path)
        try:
            record_data_for_student(marks, roster_file_name, student_numbers_in_roster)
        except:
            print(f"could not record the marks of {file}")
