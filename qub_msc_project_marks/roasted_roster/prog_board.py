import os
import openpyxl as xl
import roasted_roster.utilities as util
import datetime


def load_subject_list(prog_board_file_name, worksheet_name="Main"):
    """
    Creates a list with the subjects.

    :param prog_board_file_name: target file
    :param worksheet_name: worksheet where the data are.
    If not specified the default value is "Main"
    :return: a list with all subjects as string
    """
    wb_target = xl.load_workbook(prog_board_file_name)
    subject_list = []
    ws = wb_target[worksheet_name]
    offset = 4
    for i in range(8):
        subject = ws.cell(row=5, column=offset+2*i).value
        subject_list += [subject]
    return subject_list


def copy_from_marksheet_to_prog_board(marksheet_file_name,
                                      prog_board_file_name,
                                      student_id_list,
                                      subjects):
    prog_board_wb = xl.load_workbook(prog_board_file_name)
    prog_board_ws = prog_board_wb["Main"]
    marksheet_wb = xl.load_workbook(marksheet_file_name, data_only=True)
    marksheet_ws = marksheet_wb["SubjectBoard"]
    module_code = marksheet_ws["C1"].value
    module_index = util.find_in_list(subjects, module_code)
    date = datetime.datetime.now()
    prog_board_ws.cell(row=3, column=2).value = f"Date: {date.strftime('%d %b %G')}"
    if module_index is None:
        raise Exception("wtf")
    i = 0
    row_offset_pb = 8
    column_offset_pb = 4
    while True:
        id = marksheet_ws.cell(row=5+i, column=1).value
        name = marksheet_ws.cell(row=5+i, column=2).value
        mark = marksheet_ws.cell(row=5+i, column=3).value
        result = marksheet_ws.cell(row=5+i, column=4).value
        if result == "PH [Pass with Honours Restriction]":
            result = "PH"
        if result == "P [Pass]":
            result = "P"
        if result == "F [Fail]":
            result = "F"
        if result == "ABSM [Absent Mitigation]":
            result = "ABSM"
        if result == "MNA [Mark Not Available]":
            result = "MNA"
        i += 1
        if id is None:
            break
        position_in_list = util.find_in_list(student_id_list, id)
        is_new_student = False
        if position_in_list is None:
            is_new_student = True
            student_id_list += [id]
            position_in_list = len(student_id_list) - 1

        row = row_offset_pb + position_in_list
        column = column_offset_pb + 2 * module_index

        if is_new_student:
            prog_board_ws.cell(row=row, column=2).value = id
            prog_board_ws.cell(row=row, column=3).value = name

        prog_board_ws.cell(row=row, column=column).value = mark
        prog_board_ws.cell(row=row, column=column+1).value = result
        # ----- end of while loop
    prog_board_wb.save(prog_board_file_name)
    return student_id_list


def fill_prog_board(prog_board_file_name,
                    subject_folder_name):
    subject_folder = os.listdir(subject_folder_name)
    student_id_list = []
    subjects = load_subject_list(prog_board_file_name)
    for file in subject_folder:
        file_path = os.path.join(subject_folder_name, file)
        student_id_list = copy_from_marksheet_to_prog_board(
            file_path,
            prog_board_file_name,
            student_id_list,
            subjects)





